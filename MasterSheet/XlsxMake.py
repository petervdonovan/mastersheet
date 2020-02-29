from openpyxl import Workbook
from openpyxl import load_workbook
import re

class XlsxMake:
    '''
    Use headers and templates (existing XLSX files) to create a workbook that is 
    structured to contain all of the SOC data.
    '''
    def __init__(self):
        # Initialize with only empty default values
        self.schools = {}
        self.pathwaysTemplate = None
        self.clubsTemplate = None
        self.athleticsTemplate = None
        self.pathwaysHeaders = None
        self.clubsHeaders = None
        self.athleticsHeaders = None
        self.wb = None
        self.pathwaysTemplateType = ''
        # Load templates
        self.getTemplate()
        # Open the workbook that will be edited/created by this program
        self.openFile()
    
    def getTemplate(self):
        '''
        Load the template and header for each of the 3 possible types of sheet
        '''
        self.pathwaysTemplate = load_workbook('templates/pathways.xlsx').active
        self.pathwaysHeaders = load_workbook('headers/pathways.xlsx').active

        self.clubsTemplate = load_workbook('templates/clubs.xlsx').active
        self.clubsHeaders = load_workbook('headers/clubs.xlsx').active

        self.athleticsTemplate = load_workbook('templates/athletics.xlsx').active
        self.athleticsHeaders = load_workbook('headers/athletics.xlsx').active

    def openFile(self):
        '''
        Open the file that will be edited or created and store it as a member var.
        '''
        self.fileName = input('Workbook to edit: ')
        # Attempt to open an existing workbook for the user to edit
        try:
            self.wb = load_workbook('sheets/' + self.fileName + ".xlsx")
        except FileNotFoundError:
            # Create a new workbook.
            print('Did not find workbook of that name. Creating new workbook.')
            self.wb = Workbook()
            self.wb.remove(self.wb['Sheet'])
            print('new workbook created')

    def getUserInput(self):
        '''
        Loads all user input into the schools dictionary
        '''
        response = input("Sheet name: ")
        #An empty string response results in no save and no recursive request for another response
        if response != '':
            self.schools[response] = (self.getUrl(), self.getProgramCount())
            self.getUserInput() #Request another response

    def getProgramCount(self):
        '''
        Take user input for the number of programs at a school. If the user puts in a number that
        is too high, then that is no problem: it just means that there exists the capability to 
        load more programs, because the master spreadsheet will "look" for information in the spaces
        where those extra programs would appear, if they did exist.
        '''
        programCount = input("Maximum possible number of programs at this school: ")
        #Check to make sure the entered string can be easily evaluated as an integer
        for character in programCount:
            if not character.isdigit():
                print("Your response was not a number. ")
                return self.getProgramCount() #Request another response
        return int(programCount)

    def getUrl(self):
        '''
        Use user input to get the URL of the Google sheet from which data must be loaded.
        '''
        url = input("URL to substitute in: ")
        #Check to make sure this is a valid URL
        if not url.startswith('https://'):
            print("Your response was not a valid URL. ")
            return self.getUrl() #Request another response because previous response was not acceptable
        return url

    def replaceUrls(self, ws, url, minRow, maxRow, minColumn, maxColumn):
        '''
        Replace all instances of the string "URL" with the desired URL so that links between sheets will work
        '''
        ret = Workbook()
        retws = ret.active #get default worksheet
        #iterate over all cells in the base worksheet, ws, and copy them with modification into the worksheet to be returned, retws
        for colIndex in range(minColumn, maxColumn + 2):
            for rowIndex in range(minRow, maxRow + 2):
                if ws.cell(row = rowIndex, column = colIndex).value:
                    #copy with instances of "URL" replaced
                    try:
                        retws.cell(row = rowIndex, column = colIndex).value = ws.cell(row = rowIndex, column = colIndex).value.replace('URL', '"' + url + '"')
                    except:
                        pass
        return ret.active
    def addToNums(self, ws, numToAdd, minRow, maxRow, minCol, maxCol):
        '''
        Search through all of the numbers in the template that are surrounded by {{  }} and 
        replace each of those numbers with that number plus a given number.

        This is needed for the Pathways sheets, which contain sections of a fixed length that 
        repeat over and over in the same sheet. This means that the same type of information
        appears again and again in rows that are a known number of rows below.
        '''
        ret = Workbook()
        retws = ret.active #get default worksheet
        #iterate over all cells in the base worksheet ws
        for colIndex in range(minCol, maxCol + 2):
            for rowIndex in range(minRow, maxRow + 2):
                if ws.cell(row = rowIndex, column = colIndex).value:
                    #separate the value of the cell in ws at {{ or }}
                    parts = re.split(r'{{|}}', ws.cell(row = rowIndex, column = colIndex).value)
                    #iterate over the parts of the contents of the cell that need to be changed
                    for i in range(1, len(parts), 2):
                        parts[i] = str(int(parts[i]) + numToAdd)
                    retws.cell(row = rowIndex, column = colIndex).value = ''.join(parts)
        return ret.active
    def createSheet(self, sheetName):
        '''
        Add a new sheet wo the workbook being edited. If there already is a sheet of the same name, 
        then replace it.
        '''
        if(sheetName in self.wb.sheetnames): #it is necessary to replace a sheet that already exists
            if (len(self.wb.sheetnames) > 1):
                self.wb.remove_sheet(self.wb[sheetName])
                self.wb.create_sheet(title = sheetName)
            else: #it is not allowed to have only one sheet, so a temporary one must be created
                self.wb.create_sheet(title = 'temporary sheet')
                self.wb.remove(self.wb[sheetName])
                self.wb.create_sheet(title = sheetName)
                self.wb.remove(self.wb['temporary sheet'])
        else: self.wb.create_sheet(title = sheetName)
        return self.wb[sheetName]
    @staticmethod
    def getNumberOfCols(sheet):
        '''
        Get the number of columns based on the number manually entered into the template 
        spreadsheet, in cell D1
        '''
        return int(sheet.cell(row = 1, column = 4).value)
    @staticmethod
    def getNumberOfRows(sheet):
        '''
        Get the number of columns based on the number manually entered into the template 
        spreadsheet, in cell B1
        '''
        return int(sheet.cell(row = 1, column = 2).value)
    def makeSheetPathways(self):
        '''
        Add sheets in the main workbook for the pathways of each school.
        '''
        # Get sizes of the template and header, as manually entered at the top of the template and header
        templateNumberOfRows = XlsxMake.getNumberOfRows(self.pathwaysTemplate)
        templateNumberOfCols = XlsxMake.getNumberOfCols(self.pathwaysTemplate)
        headersNumberOfCols = XlsxMake.getNumberOfCols(self.pathwaysHeaders)
        for school in self.schools:
            # Create a worksheet with the same name as the school
            ws = self.createSheet(school)
            # Get a version of the template that is generally applicable for this school
            schoolTemplate = self.replaceUrls(self.pathwaysTemplate, self.schools[school][0], 1, templateNumberOfRows, 1, templateNumberOfCols)
            for programNumber in range(self.schools[school][1]):
                # Get a version of the template to be used for this specific program
                template = self.addToNums(schoolTemplate, programNumber * (templateNumberOfRows - self.pathwaysTemplate.min_row + 1), 1, templateNumberOfRows, 1, templateNumberOfCols)
                # Copy over the template
                for rowIndex in range(1, templateNumberOfRows + 1):
                    for colIndex in range(1, templateNumberOfCols + 1):
                        ws.cell(row = rowIndex + programNumber * (templateNumberOfRows), column = colIndex).value = template.cell(row = rowIndex + 1, column = colIndex).value
            # insert an empty row at the top for the headers (this shifts everything else down)
            ws.insert_rows(1)
            # copy over the headers from the pathway headers sheet.
            for column in range(1, headersNumberOfCols + 1):
                ws.cell(row = 1, column = column).value = self.pathwaysHeaders.cell(row = 2, column = column).value
        return self.wb
    def makeSheetClubs(self):
        '''
        Add sheets in the main workbook for the clubs of each school.
        '''
        for school in self.schools:
            # Add a new sheet for this school's clubs
            ws = self.createSheet(school + ' Clubs')
            # Get the dimensions of the template and header sheets for clubs
            templateNumberOfRows = XlsxMake.getNumberOfRows(self.clubsTemplate)
            templateNumberOfCols = XlsxMake.getNumberOfCols(self.clubsTemplate)
            headerNumberOfCols = XlsxMake.getNumberOfCols(self.clubsHeaders)
            # Drop in the correct URLs so that this sheet is linked to the correct Google sheet
            clubsTemplate = self.replaceUrls(self.clubsTemplate, self.schools[school][0], 1, templateNumberOfRows, 1, templateNumberOfCols)
            # Copy over the templates into the main workbook
            for row in range(1, templateNumberOfRows + 2): # +2 because upper bound is not inclusive AND an extra row is added for the metadata (i.e., the number of rows and columns)
                for column in range(1, templateNumberOfCols + 2):
                    ws.cell(row=row, column=column).value = clubsTemplate.cell(row=row+1, column=column).value
            ws.insert_rows(1)
            for column in range(1, headerNumberOfCols + 1):
                ws.cell(row=1, column=column).value = self.clubsHeaders.cell(row=2, column=column).value
    def makeSheetAthletics(self):
        '''
        Add sheets in the main workbook for the clubs of each school.
        '''
        for school in self.schools:
            # Add a new sheet for this school's athletics
            ws = self.createSheet(school + ' Athletics')
            # Get dimensions of athletics template and header
            templateNumberOfRows = XlsxMake.getNumberOfRows(self.athleticsTemplate)
            templateNumberOfCols = XlsxMake.getNumberOfCols(self.athleticsTemplate)
            headerNumberOfCols = XlsxMake.getNumberOfCols(self.athleticsHeaders)
            # Replace URLs with links to the correct school's Google sheet
            athleticsTemplate = self.replaceUrls(self.athleticsTemplate, self.schools[school][0], 1, templateNumberOfRows, 1, templateNumberOfCols)
            # Copy the template into the sheet in the main workbook
            for row in range(1, templateNumberOfRows + 2):
                for column in range(1, templateNumberOfCols + 2):
                    ws.cell(row=row, column=column).value = athleticsTemplate.cell(row=row+1, column=column).value
            # add a row at the top for headers
            ws.insert_rows(1)
            # Copy the headers from the XLSX file into the sheet in the main workbook
            for column in range(1, headerNumberOfCols + 1):
                ws.cell(row=1, column=column).value = self.athleticsHeaders.cell(row=2, column=column).value
    def makeSheet(self):
        '''
        Make all 3 sheets for all 11 schools inside the main workbook, and return the main workbook.
        '''
        self.makeSheetPathways()
        self.makeSheetClubs()
        self.makeSheetAthletics()
        return self.wb
    def save(self):
        '''
        Save the main workbook as XLSX file.
        '''
        try:
            self.makeSheet().save(filename = 'sheets/' + self.fileName + ".xlsx")
        except IOError:
            input("Unable to save. Is the file open in another program?")
            self.save()